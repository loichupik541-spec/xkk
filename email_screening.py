#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动筛查脚本 - 邮箱导入版与导出版对比工具

功能：
1. 读取 input/Import.xlsx（导入版邮箱）
2. 读取 input/Export.xlsx（导出版邮箱）
3. 对比筛查：
   - 导入版中已存在于导出版的邮箱（重复）
   - 导入版中不存在于导出版的邮箱（新邮箱）
4. 生成结果到 output/时间戳/ 文件夹

使用方法：
1. 将 Import.xlsx 和 Export.xlsx 放入 input 文件夹
2. 运行: python email_screening.py
3. 结果会生成在 output/时间戳/ 文件夹中
"""

import os
import sys
import re
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    print(f"错误: 缺少必要的依赖包")
    print(f"请安装: pip install pandas openpyxl")
    print(f"详细错误: {e}")
    sys.exit(1)


def get_script_directory():
    """获取脚本所在目录（支持直接运行和打包后的情况）"""
    if getattr(sys, 'frozen', False):
        # 打包后的可执行文件
        return Path(sys.executable).parent
    else:
        # 直接运行脚本
        return Path(__file__).parent.resolve()


def extract_email_from_string(text):
    """从字符串中提取邮箱地址"""
    if pd.isna(text) or text is None:
        return None
    
    text = str(text).strip()
    if not text:
        return None
    
    # 标准邮箱正则
    email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    match = re.search(email_pattern, text)
    
    if match:
        return match.group(0).lower()
    return None


def find_email_column(df):
    """自动识别包含邮箱的列"""
    email_columns = []
    
    for col in df.columns:
        # 检查列名是否包含邮箱相关关键词
        col_str = str(col).lower()
        if any(keyword in col_str for keyword in ['email', '邮箱', 'mail', 'e-mail', '电子邮件']):
            email_columns.append(col)
            continue
        
        # 检查列内容是否包含邮箱格式
        sample_values = df[col].dropna().head(10).astype(str)
        email_count = 0
        for val in sample_values:
            if extract_email_from_string(val):
                email_count += 1
        
        # 如果超过50%是邮箱格式，认为是邮箱列
        if email_count >= min(5, len(sample_values)):
            email_columns.append(col)
    
    return email_columns


def read_excel_file(filepath, file_description):
    """读取Excel文件并提取邮箱列表"""
    print(f"\n[读取] {file_description}")
    print(f"  路径: {filepath}")
    
    try:
        # 尝试读取Excel
        df = pd.read_excel(filepath, engine='openpyxl')
    except Exception as e:
        print(f"  ❌ 错误: 无法读取文件 - {e}")
        return None, [], []
    
    if df.empty:
        print(f"  ⚠️ 警告: 文件为空")
        return df, [], []
    
    print(f"  行数: {len(df)}")
    print(f"  列名: {list(df.columns)}")
    
    # 自动识别邮箱列
    email_columns = find_email_column(df)
    
    if not email_columns:
        print(f"  ⚠️ 警告: 未识别到邮箱列，尝试在所有列中搜索...")
        # 遍历所有列查找邮箱
        for col in df.columns:
            emails_found = []
            for idx, val in df[col].items():
                email = extract_email_from_string(val)
                if email:
                    emails_found.append((idx, email, col))
            if emails_found:
                email_columns = [col]
                print(f"  ✅ 在列 '{col}' 中找到 {len(emails_found)} 个邮箱")
                break
    else:
        print(f"  ✅ 识别到邮箱列: {email_columns}")
    
    if not email_columns:
        print(f"  ❌ 错误: 未找到任何邮箱")
        return df, [], []
    
    # 提取所有邮箱
    all_emails = []
    error_emails = []  # 记录格式错误的邮箱
    
    for col in email_columns:
        for idx, val in df[col].items():
            email = extract_email_from_string(val)
            if email:
                all_emails.append({
                    'row_index': idx,
                    'email': email,
                    'original_value': val,
                    'column': col
                })
            elif pd.notna(val) and str(val).strip():
                # 有内容但不是有效邮箱
                error_emails.append({
                    'row_index': idx,
                    'original_value': val,
                    'column': col,
                    'reason': '格式无效'
                })
    
    print(f"  ✅ 成功提取 {len(all_emails)} 个有效邮箱")
    if error_emails:
        print(f"  ⚠️ 发现 {len(error_emails)} 个格式错误的邮箱")
    
    return df, all_emails, error_emails


def create_result_workbook(data, sheet_name, headers):
    """创建结果工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    return wb


def save_error_emails(error_list, result_dir):
    """保存错误邮箱列表"""
    if not error_list:
        return None
    
    filepath = result_dir / "错误邮箱列表.xlsx"
    
    headers = ["来源", "行号", "原始值", "错误原因"]
    data = [[e.get('source', ''), e['row_index'] + 1, e['original_value'], e['reason']] for e in error_list]
    
    wb = create_result_workbook(data, "错误邮箱", headers)
    wb.save(filepath)
    
    return filepath


def main():
    print("=" * 70)
    print("自动筛查脚本 - 邮箱导入版与导出版对比工具")
    print("=" * 70)
    
    # 获取脚本所在目录
    script_dir = get_script_directory()
    print(f"\n脚本目录: {script_dir}")
    
    # 定义路径
    input_dir = script_dir / "input"
    output_dir = script_dir / "output"
    
    import_file = input_dir / "Import.xlsx"
    export_file = input_dir / "Export.xlsx"
    
    # 检查 input 文件夹是否存在
    if not input_dir.exists():
        print(f"\n❌ 错误: input 文件夹不存在")
        print(f"   请创建: {input_dir}")
        input("按回车键退出...")
        sys.exit(1)
    
    # 检查 output 文件夹是否存在，不存在则创建
    if not output_dir.exists():
        print(f"\n[创建] output 文件夹")
        output_dir.mkdir(parents=True, exist_ok=True)
    
    # 检查必要文件是否存在
    print(f"\n[检查] 必要文件")
    
    missing_files = []
    if not import_file.exists():
        missing_files.append(f"Import.xlsx (导入版)")
    if not export_file.exists():
        missing_files.append(f"Export.xlsx (导出版)")
    
    if missing_files:
        print(f"  ❌ 缺少以下文件:")
        for f in missing_files:
            print(f"     - {f}")
        print(f"\n  请将文件放入: {input_dir}")
        input("按回车键退出...")
        sys.exit(1)
    
    print(f"  ✅ Import.xlsx 存在")
    print(f"  ✅ Export.xlsx 存在")
    
    # 读取导入版
    import_df, import_emails, import_errors = read_excel_file(import_file, "导入版 (Import.xlsx)")
    if import_emails is None:
        input("按回车键退出...")
        sys.exit(1)
    
    # 读取导出版
    export_df, export_emails, export_errors = read_excel_file(export_file, "导出版 (Export.xlsx)")
    if export_emails is None:
        input("按回车键退出...")
        sys.exit(1)
    
    # 提取邮箱集合用于对比
    import_email_set = set(e['email'] for e in import_emails)
    export_email_set = set(e['email'] for e in export_emails)
    
    print("\n" + "=" * 70)
    print("筛查结果")
    print("=" * 70)
    
    # 分类
    duplicate_emails = import_email_set & export_email_set  # 已存在
    new_emails = import_email_set - export_email_set  # 新邮箱
    
    print(f"\n导入版邮箱总数: {len(import_email_set)}")
    print(f"导出版邮箱总数: {len(export_email_set)}")
    print(f"\n分类结果:")
    print(f"  🔴 已存在于导出版 (重复): {len(duplicate_emails)} 个")
    print(f"  🟢 不存在于导出版 (新邮箱): {len(new_emails)} 个")
    
    # 生成时间戳
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 准备输出数据
    print("\n" + "=" * 70)
    print("生成结果文件")
    print("=" * 70)
    
    # 创建以时间命名的子文件夹
    result_dir = output_dir / timestamp
    result_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n  📁 结果文件夹: {result_dir.name}/")
    
    # 1. 生成重复邮箱列表
    if duplicate_emails:
        duplicate_data = []
        for email in sorted(duplicate_emails):
            import_info = next((e for e in import_emails if e['email'] == email), None)
            duplicate_data.append([
                email,
                import_info['row_index'] + 1 if import_info else ''
            ])
        
        filepath = result_dir / "重复邮箱.xlsx"
        wb = create_result_workbook(
            duplicate_data,
            "重复邮箱",
            ["邮箱", "行号"]
        )
        wb.save(filepath)
        print(f"    📄 重复邮箱.xlsx ({len(duplicate_data)} 条)")
    else:
        print(f"    ℹ️ 无重复邮箱")
    
    # 2. 生成新邮箱列表
    if new_emails:
        new_data = []
        for email in sorted(new_emails):
            import_info = next((e for e in import_emails if e['email'] == email), None)
            new_data.append([
                email,
                import_info['row_index'] + 1 if import_info else ''
            ])
        
        filepath = result_dir / "新邮箱.xlsx"
        wb = create_result_workbook(
            new_data,
            "新邮箱",
            ["邮箱", "行号"]
        )
        wb.save(filepath)
        print(f"    📄 新邮箱.xlsx ({len(new_data)} 条)")
    else:
        print(f"    ℹ️ 无新邮箱")
    
    # 3. 生成汇总报告
    summary_data = [
        ["导入版邮箱总数", len(import_email_set)],
        ["导出版邮箱总数", len(export_email_set)],
        ["", ""],
        ["重复邮箱数", len(duplicate_emails)],
        ["新邮箱数", len(new_emails)],
        ["", ""],
        ["导入版错误邮箱", len(import_errors)],
        ["导出版错误邮箱", len(export_errors)],
    ]
    
    filepath = result_dir / "筛查汇总.xlsx"
    wb = create_result_workbook(
        summary_data,
        "汇总",
        ["项目", "数量"]
    )
    wb.save(filepath)
    print(f"    📄 筛查汇总.xlsx")
    
    # 4. 保存错误邮箱列表
    all_errors = import_errors + export_errors
    if all_errors:
        for e in import_errors:
            e['source'] = '导入版'
        for e in export_errors:
            e['source'] = '导出版'
        
        error_filepath = save_error_emails(all_errors, result_dir)
        if error_filepath:
            print(f"    📄 错误邮箱列表.xlsx ({len(all_errors)} 条)")
    
    print("\n" + "=" * 70)
    print(f"✅ 完成！结果已保存到: {result_dir}")
    print("=" * 70)
    
    input("\n按回车键退出...")


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️ 用户中断")
        sys.exit(0)
    except Exception as e:
        print(f"\n\n❌ 程序错误: {e}")
        import traceback
        traceback.print_exc()
        input("\n按回车键退出...")
        sys.exit(1)
